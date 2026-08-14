"""Xlsx generator for xlsx output."""

import asyncio
from datetime import datetime
import json
import logging
from pathlib import Path
from typing import Any, Dict, List

from tempfile import NamedTemporaryFile
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment

from .report_renderer import ReportRenderer
from .repository import Report
from .template_renderer import template_renderer

logger = logging.getLogger(__name__)

class XlsxRenderer(ReportRenderer):
    """Generate XLSX using Xlsx from raw data."""

    def __init__(self):
        self.logoname = "thermoelectrica_logo.png"
        self.file_xlsx = "act_template.xlsx"
        self.plant_reference = "plant_reference.json"
        self.service_type = {"Монтаж": "montage", "Осмотр": "inspection"}
        
    @property
    def format_name(self) -> str:
        return "xlsx"

    @property
    def file_extension(self) -> str:
        return "xlsx"

    async def render_preview(
        self,
        report: Report,
        parameters: Dict[str, Any],
        query_results: Dict[str, List[Dict[str, Any]]],
    ) -> str | None:
        """Отрендерить HTML-превью. Возвращает None, если формат не поддерживает превью."""
        return template_renderer.render(report, parameters, query_results)

    @property
    def supports_preview(self) -> bool:
        return False
    
    @classmethod
    def montage_result(cls, count: int) -> str:
        return f"Термоиндикаторы установлены в количестве {count} шт."
    
    @classmethod
    def inspection_result(cls, protocol: str) -> str:
        date = ""
        try:
            number, date = protocol.split(" от ")
            dt = datetime.strptime(date, "%d.%m.%Y")
            dt_formatted = template_renderer.add_day_parser(dt, 0)
            return (
                "Термоиндикаторы функционируют исправно. "
                f"Температурное состояние электрооборудования "
                f"определено (протокол №{number} от {dt_formatted})."
            )
        except Exception:
            if not date:
                raise ValueError("Дата протокола не введена или нарушен формат ввода")
            raise ValueError(f"Нарушен формат ввода даты: {date}")
    
    def get_plant_content(
            self, 
            plant: str, 
            control_type: str | None = None, 
            data: str | None = None,
            json_file: Path | str = Path("/"),
    ) -> tuple[str, Any, list[dict]]:
        # Загрузка справочника станций
        with open(json_file, "r", encoding="utf-8") as file:
            PLANT = json.load(file)
        for item in PLANT:
            if item["plant_name"] == plant:
                return (
                    item[control_type][f"{control_type}_name"],
                    getattr(XlsxRenderer, item[control_type][f"{control_type}_result"])(data),
                    item["signatories"],
                )
        raise ValueError(f"{plant} отсутствует в справочниках")
        
    def render_xlsx(self, report: Report, params: Dict[str, Any]) -> bytes:
        # Проверяем наличие файлов XLSX и JSON 
        xlsx_file = Path(report.path) / self.file_xlsx
        json_file = Path(report.path) / self.plant_reference

        for path, name in [(xlsx_file, self.file_xlsx), (json_file, self.plant_reference)]:
            if not path.exists():
                raise ValueError(f"Отсутствует {name} в {report.path}")

        if not self.query_results.get("data"):
                raise ValueError(f"По {params['plant_name']} нет данных для формирования документа")
        
        try:
            # Открываем шаблон
            wb = load_workbook(xlsx_file)
    
            # Получаем активный лист
            sheet = wb.active
            # Называем лист
            sheet.title = "Технический акт"

            # Вставляем нумерацию
            sheet.oddFooter.right.text = "Страница &P из &N"
            sheet.evenFooter.right.text = "Страница &P из &N"
            
            # Задаем стили
            sd = Side(style="thin")
            border=Border(left=sd, top=sd, right=sd, bottom=sd)
            font=Font(name="Times New Roman", bold=False, size=14, italic=False)
            alignment=Alignment(
                wrap_text=True,
                horizontal="left",
                vertical="center",
            )
            base_cell = NamedStyle(
                name="base_cell", 
                font=font,
                alignment=Alignment(
                    wrap_text=True, 
                    horizontal="center", 
                    vertical="center",
                ),
            )
            table_cell = NamedStyle(
                name="table_cell", 
                font=font,
                alignment=alignment,
                border=border
            )
            underline_bold = NamedStyle(
                name="underline_bold",
                font=Font(
                    name="Times New Roman", 
                    bold=True, 
                    size=14, 
                    italic=False, 
                    underline="single"
                ),
            )
            notice_cell = NamedStyle(
                name="notice_cell",
                font=Font(
                    name="Times New Roman", 
                    bold=True, 
                    size=14, 
                    italic=False, 
                )
            )

            # Регистрируем стили
            for style in (base_cell, table_cell, underline_bold, notice_cell):
                if style.name not in wb.named_styles:
                    wb.add_named_style(style)

            # Заголовок документа
            sheet.cell(
                row=1, 
                column=1, 
                value=f"ТЕХНИЧЕСКИЙ АКТ № {params['act_number']}"
            )

            # Дата формирования акта
            cell = sheet.cell(
                row=2, 
                column=5, 
                value=f"{template_renderer.add_day_parser(params['period_end'], 0)}"
            )
            cell.style = base_cell
            cell.alignment = Alignment(horizontal="right")

            # Информация о сервисном обслуживании   
            cell = sheet.cell(
                row=4, 
                column=1, 
                value=(
                    "по сервисному обслуживанию систем контроля температурного "
                    "состояния контактов и контактных соединений электрооборудования"
                )
            )
            cell.font = Font(name="Times New Roman", size=16, italic=True)
            
            # Определяем стартовую строку
            start_row = sheet.max_row

            signatories = []

            # Формируем таблицу
            for row_idx, row_data in enumerate(self.query_results["data"], start=start_row):
                # Заполняем строку: индекс row_idx соответствует next row
                sheet.row_dimensions[row_idx].height = 60
                for col_idx in range(1, 6):
                    # установка значений
                    if col_idx == 1:
                        cell_obj = sheet.cell(row=row_idx, column=col_idx, value=(row_idx - start_row + 1))
                    if col_idx == 2:
                        cell_obj = sheet.cell(row=row_idx, column=col_idx, value=params["plant_name"])
                    if col_idx == 3:
                        cell_obj = sheet.cell(
                            row=row_idx, 
                            column=col_idx, 
                            value=row_data["equipment_path"]
                        )
                    
                    if col_idx == 4:
                        data = (
                            (row_data["sticker_count"] or "0")
                            if params["service_type"] == "Монтаж"
                            else params["protocol_number"]
                        ) 
                        service, result, signatories = self.get_plant_content(
                            params["plant_name"], 
                            self.service_type[params["service_type"]],
                            data,
                            json_file,
                        )
                        cell_obj = sheet.cell(row=row_idx, column=col_idx, value=service)
                    if col_idx == 5:
                        cell_obj = sheet.cell(row=row_idx, column=col_idx, value=result)
                    
                    # установка стилей
                    cell_obj.style=table_cell
                    if col_idx in [1, 2]:
                        cell_obj.alignment=Alignment(horizontal="center", vertical="center", wrap_text=True)
                    if col_idx in [2, 4, 5]:
                        cell_obj.font = Font(name="Times New Roman", bold=False, size=14, italic=True)
            
            # Добавляем пустую строку
            sheet.append([None] * sheet.max_column)

            current_row = sheet.max_row + 1
            
            # Выводим заметку
            sheet.cell(
                row=current_row, 
                column=1, 
                value=(
                    "  *  - перечисляются все Услуги, которые указаны "
                    "в Договоре по данному объекту."
                )
            ).style = notice_cell
            
            # Попытка управления разрывами
            #row_number = 20  # номер строки, перед которой вставить разрыв
            #sheet.cell(row=current_row + 1, column=2, value=sheet.max_row).style = underline_bold
            #sheet.page_setup.autoPageBreaks = True
            #sheet.page_setup.fitToPage = False
            #page_break = Break(id=sheet.max_row)  # создаём объект разрыва
            #sheet.row_breaks.append(page_break)  # добавляем разрыв в коллекцию
            #print("Горизонтальные разрывы: ", sheet.row_breaks, "MAX_ROW: ", sheet.max_row)

            sheet.append([None] * sheet.max_column)
            current_row = sheet.max_row + 1

            # Объединяем ячейки
            sheet.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
            sheet.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=5)

            # Создаем блок подписей
            sheet.cell(row=current_row, column=2, value="от ЗАКАЗЧИКА:").style = underline_bold
            sheet.cell(row=current_row, column=4, value="от ПОДРЯДЧИКА:").style = underline_bold

            current_row = sheet.max_row + 1

            provider = "Генеральный директор\nООО «ТермоЭлектрика»"
            sign_point = "_________________________/А.В. Лесив /"

            # Правая колонка подрядчика
            sheet.row_dimensions[current_row].height = 50
            cell = sheet.cell(row=current_row, column=4, value=provider)
            cell.style = notice_cell
            cell.alignment = alignment
            sheet.row_dimensions[current_row + 1].height = 30
            cell = sheet.cell(row=current_row + 1, column=4, value=sign_point)
            cell.style = notice_cell
            
            # Левая колонка заказчика
            for item in signatories:
                sheet.row_dimensions[current_row].height = 50
                sheet.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
                cell = sheet.cell(row=current_row, column=2, value=item["position"])
                cell.style = notice_cell
                cell.alignment = alignment
                sheet.row_dimensions[current_row + 1].height = 30
                sheet.merge_cells(start_row=current_row + 1, start_column=2, end_row=current_row + 1, end_column=3)
                sign_point = f"_________________________/{item['full_name']} /"
                cell = sheet.cell(row=current_row + 1, column=2, value=sign_point)
                cell.style = notice_cell
                sheet.append([None] * sheet.max_column)
                sheet.append([None] * sheet.max_column)
                current_row = sheet.max_row + 1
            
            # Сохраняем изменения
            with NamedTemporaryFile() as tmp:
                wb.save(tmp.name)
                tmp.seek(0)
                file_stream = tmp.read()

        except Exception as exc:
            logger.error(f"XLSX rendering failed: {exc}")
            raise RuntimeError(f"XLSX rendering failed: {exc}")
        return file_stream

    async def render(
        self,
        report: Report,
        parameters: Dict[str, Any],
        query_results: Dict[str, List[Dict[str, Any]]],
        base_url: str | None = None,
    ) -> bytes:
        """
        Takes Args and make XLSX using XLSX.
        Args:
            report: Report object with template
            parameters: User-provided parameters
            query_results: Query results as DataFrames
            base_url: Base URL for resolving relative paths
        Returns:
            XLSX file content
        """
        self.query_results = query_results

        try:
            logger.info(f"Generating XLSX from raw data (source: {report.path})")

            # If base_url not provided, use report.path as base
            if base_url is None:
                # Convert to absolute path first to avoid "relative paths can't be expressed as file URIs" error
                absolute_path = report.path.resolve()
                base_url = absolute_path.as_uri()

            # рендерим XLSX
            xlsx_bytes = await asyncio.to_thread(self.render_xlsx, report, parameters)
            logger.info(f"XLSX generated successfully, size: {len(xlsx_bytes)} bytes")
            return xlsx_bytes

        except Exception as e:
            logger.error(f"XLSX generation failed: {e}")
            raise RuntimeError(f"XLSX generation failed: {e}")

# Global XlsxRenderer instance
xlsx_renderer = XlsxRenderer()
